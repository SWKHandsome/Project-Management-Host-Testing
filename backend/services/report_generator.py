"""
Report generation service
Generates assessment reports in various formats
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from config import Config

class ReportGenerator:
    """Generate assessment reports"""
    
    def __init__(self):
        self.reports_dir = Config.REPORTS_DIR
    
    def generate_individual_report(self, submission, format='pdf'):
        """Generate individual report for a submission
        
        Args:
            submission: Submission data
            format: 'pdf' or 'txt' (default: 'pdf')
        """
        try:
            if format.lower() == 'pdf':
                return self.generate_pdf_report(submission)
            else:
                return self.generate_text_report(submission)
        except Exception as e:
            print(f"  ✗ Error generating report: {e}")
            raise
    
    def generate_text_report(self, submission):
        """Generate individual text report for a submission"""
        try:
            student_id = submission.get('student_id', 'UNKNOWN')
            assessment = submission.get('assessment', {})
            
            # Create report filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"report_{student_id}_{timestamp}.txt"
            filepath = os.path.join(self.reports_dir, filename)
            
            # Generate report content
            report_content = self.format_individual_report(submission, assessment)
            
            # Write to file
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(report_content)
            
            print(f"  ✓ Text report generated: {filename}")
            return filepath
            
        except Exception as e:
            print(f"  ✗ Error generating text report: {e}")
            raise
    
    def generate_pdf_report(self, submission):
        """Generate PDF report for a submission"""
        try:
            student_id = submission.get('student_id', 'UNKNOWN')
            student_name = submission.get('student_name', 'UNKNOWN')
            file_name = submission.get('file_name', 'UNKNOWN')
            submitted_at = submission.get('submitted_at', datetime.now())
            evaluated_at = submission.get('evaluated_at', datetime.now())
            
            assessment = submission.get('assessment', {})
            total_score = assessment.get('total_score', 0)
            grade = assessment.get('grade', 'F')
            breakdown = assessment.get('breakdown', {})
            feedback = assessment.get('feedback', {})
            
            # Create PDF filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"report_{student_id}_{timestamp}.pdf"
            filepath = os.path.join(self.reports_dir, filename)
            
            # Create PDF
            doc = SimpleDocTemplate(filepath, pagesize=letter,
                                   rightMargin=72, leftMargin=72,
                                   topMargin=72, bottomMargin=18)
            
            # Container for elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a5490'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#2c5aa0'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            # Title
            elements.append(Paragraph("PROGRAMMING LOGIC ASSIGNMENT", title_style))
            elements.append(Paragraph("ASSESSMENT REPORT", title_style))
            elements.append(Spacer(1, 20))
            
            # Student Information
            elements.append(Paragraph("STUDENT INFORMATION", heading_style))
            student_data = [
                ['Student ID:', student_id],
                ['Student Name:', student_name],
                ['Assignment File:', file_name],
                ['Submission Date:', submitted_at.strftime('%Y-%m-%d %H:%M:%S') if isinstance(submitted_at, datetime) else str(submitted_at)],
                ['Evaluation Date:', evaluated_at.strftime('%Y-%m-%d %H:%M:%S') if isinstance(evaluated_at, datetime) else str(evaluated_at)]
            ]
            student_table = Table(student_data, colWidths=[2*inch, 4*inch])
            student_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f0f8')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(student_table)
            elements.append(Spacer(1, 20))
            
            # Assessment Results
            elements.append(Paragraph("ASSESSMENT RESULTS", heading_style))
            grade_color = colors.green if total_score >= 70 else colors.orange if total_score >= 50 else colors.red
            results_data = [
                ['Overall Score:', f"{total_score}/100"],
                ['Grade:', grade]
            ]
            results_table = Table(results_data, colWidths=[2*inch, 4*inch])
            results_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f0f8')),
                ('BACKGROUND', (1, 1), (1, 1), grade_color),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('TEXTCOLOR', (1, 1), (1, 1), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(results_table)
            elements.append(Spacer(1, 20))
            
            # Detailed Breakdown
            elements.append(Paragraph("DETAILED BREAKDOWN", heading_style))
            breakdown_data = [['Category', 'Score', 'Max Score', 'Percentage']]
            for category, data in breakdown.items():
                category_name = category.replace('_', ' ').title()
                score = data.get('score', 0)
                max_score = data.get('max_score', 0)
                percentage = data.get('percentage', 0)
                breakdown_data.append([category_name, f"{score:.1f}", f"{max_score}", f"{percentage:.1f}%"])
            
            breakdown_table = Table(breakdown_data, colWidths=[2.5*inch, 1*inch, 1.2*inch, 1.3*inch])
            breakdown_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5aa0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')])
            ]))
            elements.append(breakdown_table)
            elements.append(Spacer(1, 20))
            
            # Feedback
            strengths = assessment.get('strengths', [])
            improvements = assessment.get('improvements', [])
            recommendations = assessment.get('recommendations', [])
            
            if strengths:
                elements.append(Paragraph("STRENGTHS", heading_style))
                for strength in strengths:
                    elements.append(Paragraph(f"✓ {strength}", styles['Normal']))
                elements.append(Spacer(1, 12))
            
            if improvements:
                elements.append(Paragraph("AREAS FOR IMPROVEMENT", heading_style))
                for improvement in improvements:
                    elements.append(Paragraph(f"○ {improvement}", styles['Normal']))
                elements.append(Spacer(1, 12))
            
            if recommendations:
                elements.append(Paragraph("RECOMMENDATIONS", heading_style))
                for recommendation in recommendations:
                    elements.append(Paragraph(f"→ {recommendation}", styles['Normal']))
            
            # Build PDF
            doc.build(elements)
            
            print(f"  ✓ PDF report generated: {filename}")
            return filepath
            
        except Exception as e:
            print(f"  ✗ Error generating PDF report: {e}")
            raise
    
    def format_individual_report(self, submission, assessment):
        """Format individual report content"""
        student_id = submission.get('student_id', 'UNKNOWN')
        student_name = submission.get('student_name', 'UNKNOWN')
        file_name = submission.get('file_name', 'UNKNOWN')
        submitted_at = submission.get('submitted_at', datetime.now())
        evaluated_at = submission.get('evaluated_at', datetime.now())
        
        total_score = assessment.get('total_score', 0)
        grade = assessment.get('grade', 'F')
        breakdown = assessment.get('breakdown', {})
        feedback = assessment.get('feedback', {})
        
        report = f"""
{"="*80}
                    PROGRAMMING LOGIC ASSIGNMENT
                        ASSESSMENT REPORT
{"="*80}

STUDENT INFORMATION
{"-"*80}
Student ID       : {student_id}
Student Name     : {student_name}
Assignment File  : {file_name}
Submission Date  : {submitted_at.strftime('%Y-%m-%d %H:%M:%S') if isinstance(submitted_at, datetime) else submitted_at}
Evaluation Date  : {evaluated_at.strftime('%Y-%m-%d %H:%M:%S') if isinstance(evaluated_at, datetime) else evaluated_at}

{"="*80}
                         ASSESSMENT RESULTS
{"="*80}

OVERALL SCORE: {total_score}/100    GRADE: {grade}

{"="*80}
                        DETAILED BREAKDOWN
{"="*80}
"""
        
        # Add breakdown for each category
        for category, data in breakdown.items():
            category_name = category.replace('_', ' ').title()
            score = data.get('score', 0)
            max_score = data.get('max_score', 0)
            percentage = data.get('percentage', 0)
            category_feedback = data.get('feedback', [])
            
            report += f"\n{category_name.upper()}\n"
            report += f"{'-'*80}\n"
            report += f"Score: {score}/{max_score} ({percentage}%)\n"
            
            if category_feedback:
                report += f"Feedback:\n"
                for item in category_feedback:
                    report += f"  • {item}\n"
            report += "\n"
        
        # Add overall feedback
        report += f"{'='*80}\n"
        report += f"                      OVERALL FEEDBACK\n"
        report += f"{'='*80}\n\n"
        
        strengths = feedback.get('strengths', [])
        improvements = feedback.get('improvements', [])
        recommendations = feedback.get('recommendations', [])
        
        if strengths:
            report += f"STRENGTHS:\n"
            for strength in strengths:
                report += f"  ✓ {strength}\n"
            report += "\n"
        
        if improvements:
            report += f"AREAS FOR IMPROVEMENT:\n"
            for improvement in improvements:
                report += f"  ○ {improvement}\n"
            report += "\n"
        
        if recommendations:
            report += f"RECOMMENDATIONS:\n"
            for recommendation in recommendations:
                report += f"  → {recommendation}\n"
            report += "\n"
        
        report += f"{'='*80}\n"
        report += f"                    END OF REPORT\n"
        report += f"{'='*80}\n"
        
        return report
    
    def generate_spreadsheet(self, submissions):
        """Generate Excel spreadsheet with all assessments"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"assessments_summary_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Assessments"
            
            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                'Student ID', 'Student Name', 'File Name', 'Submission Date',
                'Total Score', 'Grade', 
                'Logic Design', 'Flowchart', 'Pseudocode', 'Formatting', 'Documentation',
                'Status'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data rows
            for row_idx, submission in enumerate(submissions, 2):
                assessment = submission.get('assessment', {})
                breakdown = assessment.get('breakdown', {})
                
                # Extract scores
                logic_score = breakdown.get('logic_design', {}).get('score', 0)
                flowchart_score = breakdown.get('flowchart', {}).get('score', 0)
                pseudocode_score = breakdown.get('pseudocode', {}).get('score', 0)
                formatting_score = breakdown.get('formatting', {}).get('score', 0)
                documentation_score = breakdown.get('documentation', {}).get('score', 0)
                
                row_data = [
                    submission.get('student_id', 'N/A'),
                    submission.get('student_name', 'N/A'),
                    submission.get('file_name', 'N/A'),
                    submission.get('submitted_at', datetime.now()).strftime('%Y-%m-%d') if isinstance(submission.get('submitted_at'), datetime) else 'N/A',
                    assessment.get('total_score', 0),
                    assessment.get('grade', 'F'),
                    logic_score,
                    flowchart_score,
                    pseudocode_score,
                    formatting_score,
                    documentation_score,
                    submission.get('status', 'N/A')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Color code grades
                    if col == 6:  # Grade column
                        if value in ['A+', 'A', 'A-']:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value in ['B+', 'B', 'B-']:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif value in ['C+', 'C', 'C-', 'D']:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
            
            # Adjust column widths
            column_widths = [15, 20, 30, 15, 12, 8, 12, 12, 12, 12, 15, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width
            
            # Add statistics sheet
            stats_ws = wb.create_sheet("Statistics")
            self.add_statistics_sheet(stats_ws, submissions)
            
            # Save workbook
            wb.save(filepath)
            
            print(f"  ✓ Spreadsheet generated: {filename}")
            return filepath
            
        except Exception as e:
            print(f"  ✗ Error generating spreadsheet: {e}")
            raise
    
    def add_statistics_sheet(self, ws, submissions):
        """Add statistics to spreadsheet"""
        try:
            # Title
            ws['A1'] = 'ASSESSMENT STATISTICS'
            ws['A1'].font = Font(bold=True, size=14)
            
            # Calculate statistics
            total = len(submissions)
            evaluated = len([s for s in submissions if s.get('status') == 'evaluated'])
            
            scores = [s.get('assessment', {}).get('total_score', 0) for s in submissions if s.get('status') == 'evaluated']
            avg_score = sum(scores) / len(scores) if scores else 0
            max_score = max(scores) if scores else 0
            min_score = min(scores) if scores else 0
            
            passed = len([s for s in scores if s >= Config.PASS_THRESHOLD])
            failed = len(scores) - passed
            
            # Add data
            stats = [
                ('', ''),
                ('Total Submissions:', total),
                ('Evaluated:', evaluated),
                ('Pending:', total - evaluated),
                ('', ''),
                ('Average Score:', f'{avg_score:.2f}'),
                ('Highest Score:', max_score),
                ('Lowest Score:', min_score),
                ('', ''),
                ('Passed:', passed),
                ('Failed:', failed),
                ('Pass Rate:', f'{(passed/evaluated*100):.2f}%' if evaluated > 0 else '0%')
            ]
            
            for row, (label, value) in enumerate(stats, 3):
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            
        except Exception as e:
            print(f"  ⚠ Error adding statistics: {e}")
